"""렛유인 KDC 강의계획서 xlsx 생성 — 양식 동일 형식"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── 공통 스타일 헬퍼 ────────────────────────────────────────────
THIN = Side(style="thin")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def apply(cell, value=None, bold=False, color="FF000000", bg=None,
          halign="left", valign="center", size=10, wrap=True):
    if value is not None:
        cell.value = value
    cell.font = Font(bold=bold, color=color, size=size, name="맑은 고딕")
    if bg:
        cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=wrap)
    cell.border = THIN_BORDER


# ── 워크북 생성 ─────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "강의계획서"
ws.sheet_view.showGridLines = False

# 컬럼 너비 — 양식과 동일
ws.column_dimensions["A"].width = 18.0
ws.column_dimensions["B"].width = 8.0
ws.column_dimensions["C"].width = 33.5
ws.column_dimensions["D"].width = 64.875
ws.column_dimensions["E"].width = 12.0


# ══════════════════════════════════════════════════════════════
# ■ 헤더 영역 (행 1~6)
# ══════════════════════════════════════════════════════════════

# 행 1 — 제목
ws.row_dimensions[1].height = 30.0
ws.merge_cells("A1:E1")
apply(ws["A1"], "강의 계획서",
      bold=True, color="FFFFFFFF", bg="FF1F4E79",
      halign="center", size=14)

# 행 2~5 — 기본 정보 (A=레이블 bg=BDD7EE, B:E 병합=값)
info_rows = [
    (2, "강의명",
     "디스플레이 취업준비생을 위한 AI 바이브 코딩 실전"),
    (3, "강의분량",
     "총 16강 (1~14강·16강 각 30분 / 15강 60분) = 총 약 8시간 30분"),
    (4, "수강 대상",
     "디스플레이 관련 학과 대학교 3~4학년 / 취업준비생 (주 대상) | "
     "디스플레이 업계 재직자·이직 준비생 (확대 대상)"),
    (5, "강의 목표",
     "① 코드 없이 AI(Claude·Gemini)로 실제 배포 앱 제작  "
     "② GitHub 포트폴리오 완성  "
     "③ 디스플레이 도메인(공정·수율·소재) × AI 개발 접목  "
     "④ 배포 플랫폼 10종 이상 경험"),
]
for r, label, value in info_rows:
    ws.row_dimensions[r].height = 21.95
    apply(ws.cell(r, 1), label, bold=True, bg="FFBDD7EE", halign="center")
    ws.merge_cells(f"B{r}:E{r}")
    apply(ws.cell(r, 2), value, halign="left")

# 행 6 — 컬럼 헤더
ws.row_dimensions[6].height = 25.0
for c, h in enumerate(
        ["강의구분", "차시", "차시 주제", "세부 내용", "예상시간"], 1):
    apply(ws.cell(6, c), h,
          bold=True, color="FFFFFFFF", bg="FF2E75B6", halign="center")


# ══════════════════════════════════════════════════════════════
# ■ 강의 데이터
#   각 항목: (강의구분, 차시, 차시 주제, 세부 내용, 예상시간)
#   강의구분이 None이면 A열은 빈 셀 (양식 동일)
# ══════════════════════════════════════════════════════════════

PART_BG  = "FFFFFFFF"    # 일반 PART 배경 (흰색)
PROJ_BG  = "FFFFC000"    # PROJECT 행 배경 (주황-노랑)
TOPIC_HL = "FFFFFF00"    # 주요 차시 주제 강조 (노랑)

rows = [
    # ── PART 1 ─────────────────────────────────────────────
    ("PART 1\n세팅 &\n바이브 코딩\n철학",
     "1강",
     "AI 마인드셋 + Claude Pro 시작",
     "- 바이브 코딩이란? (Andrej Karpathy 선언 소개)\n"
     "- AI 개발 생태계 지형도: Claude / Gemini / GPT 비교\n"
     "- Claude의 Constitutional AI 원칙 이해\n"
     "- 실습: Claude.ai 로그인 → Projects 생성 → 첫 프롬프트",
     "30분"),
    (None,
     "2강",
     "GitHub + Claude Code 환경 세팅",
     "- Git vs GitHub 개념 (버전관리 필요성)\n"
     "- Claude Code 설치 및 초기 설정 (npm install -g @anthropic-ai/claude-code)\n"
     "- CLAUDE.md 작성법: AI에게 프로젝트 맥락 전달\n"
     "- 실습: GitHub 계정 생성 → Claude Code 설치 → 첫 commit & push",
     "30분"),

    # ── PART 2 ─────────────────────────────────────────────
    ("PART 2\nClaude Code\n심화 +\n에이전트",
     "3강",
     "Claude Code Skills & 커스텀 자동화",
     "- Skills 만들기: /commit, /review 등 커스텀 명령\n"
     "- 프롬프트 패턴: Chain-of-Thought, Few-shot 활용법\n"
     "- cokacdir 레포 탐색: AI 자연어 터미널 파일 매니저 소개\n"
     "- 실습: 커스텀 /commit 스킬 생성 및 자동화 설정",
     "30분"),
    (None,
     "4강",
     "Claude Agent 모드 + CLI-Anything",
     "- Agent 모드 동작 원리 (서브에이전트 병렬 실행)\n"
     "- CLI-Anything 레포 분석: 어떤 CLI 도구든 AI 에이전트로 전환\n"
     "- 기존 CLI 도구를 AI화하는 방법\n"
     "- 실습: CLI-Anything으로 나만의 AI 도구 만들기\n"
     "※ 프로젝트 1 과제 안내 (1주일 후 제출)",
     "30분"),

    # ── PROJECT P1 ──────────────────────────────────────────
    ("PROJECT\n[P1]",
     "P1",
     "나의 디스플레이 관심 분야 소개 웹페이지",
     "- 주제: 본인의 디스플레이 관심 분야 소개 웹페이지 제작 후 GitHub Pages 배포\n"
     "- 하: 완성 템플릿 수정 후 배포\n"
     "- 중: 관심 분야 선택 후 Claude Code로 페이지 직접 제작\n"
     "- 상: 기획~디자인~배포 전 과정 독립 수행\n"
     "- 제출: GitHub Pages 배포 URL",
     "1주일"),

    # ── PART 3 ─────────────────────────────────────────────
    ("PART 3\n데이터\n시각화\n앱 3종",
     "5강",
     "Python Streamlit 앱 제작 & 배포",
     "- Streamlit 구조 및 핵심 위젯 (st.slider, st.plotly_chart 등)\n"
     "- Claude Code로 앱 코드 자동 생성\n"
     "- Streamlit Community Cloud 무료 배포\n"
     "- 실습: 디스플레이 PPI 계산기 앱 제작 → 배포 URL 공유\n"
     "※ 오프닝 훅: GitHub Star 38만 돌파",
     "30분"),
    (None,
     "6강",
     "R Shiny 대시보드 제작 & 배포",
     "- R Shiny 기본 구조 (ui / server)\n"
     "- Streamlit vs Shiny 비교\n"
     "- Claude Code로 Shiny 코드 생성 및 shinyapps.io 배포\n"
     "- 실습: 수율 트렌드 대시보드 제작\n"
     "⚠️ 사전 준비 필수: R + RStudio + shinyapps.io 계정 설치\n"
     '   install.packages(c("shiny","shinydashboard","ggplot2","dplyr","rsconnect"))',
     "30분"),
    (None,
     "7강",
     "Quarto 공정 분석 보고서 & GitHub Pages 배포",
     "- Quarto 문서 구조 (.qmd 파일): 코드 + 설명 + 시각화 통합\n"
     "- AI 보조 공정 분석 보고서 작성\n"
     "- GitHub Pages 자동 게시 설정\n"
     "- 실습: 공정 분석 보고서 작성 → GitHub Pages 배포\n"
     "※ 오프닝 훅: PPT 수정 지옥 탈출",
     "30분"),

    # ── PART 4 ─────────────────────────────────────────────
    ("PART 4\nAI API +\n시뮬레이션",
     "8강",
     "Google AI Studio + Gemini API 챗봇",
     "- Google AI Studio에서 API Key 발급\n"
     "- Claude vs Gemini 성능 비교표\n"
     "- .env 파일로 API Key 안전 관리\n"
     "- 실습: Gemini API로 디스플레이 Q&A 챗봇 제작 (gemini_chatbot.py)\n"
     "※ 오프닝 훅: 하루 1,500회 무료",
     "30분"),
    (None,
     "9강",
     "AI 시뮬레이션 만들기",
     "- 디스플레이 시뮬레이션 아이디어: 픽셀 구조 / 색역 / 포토공정 시각화\n"
     "- Claude Code로 시뮬레이션 앱 자동 생성\n"
     "- 실습: RGB 색역 시뮬레이터 제작 → Streamlit Cloud 배포",
     "30분"),
    (None,
     "10강",
     "웹앱 배포 플랫폼 심화 (GitHub Pages / Streamlit / Firebase)",
     "- 배포 플랫폼 비교표: GitHub Pages / Streamlit Cloud / Firebase Studio\n"
     "- 각 플랫폼 장단점 및 선택 기준\n"
     "- 실습: 디스플레이 스펙 비교 웹앱을 3개 플랫폼에 배포 비교\n"
     "※ 프로젝트 2 과제 안내 (1주일 후 제출)",
     "30분"),

    # ── PROJECT P2 ──────────────────────────────────────────
    ("PROJECT\n[P2]",
     "P2",
     "디스플레이 공정 시뮬레이터 or 데이터 분석 앱",
     "- 주제: 공정 시뮬레이터 또는 데이터 분석 앱 제작 후 Streamlit/Shiny 배포\n"
     "- 하: 제공 코드 실행·배포 (SECOM 데이터셋 활용)\n"
     "- 중: 기능 추가·시각화 수정 (Streamlit or Shiny)\n"
     "- 상: 데이터 선정~모델링~시각화~배포 독립 수행\n"
     "- 제출: Streamlit 또는 Shiny 배포 URL",
     "1주일"),

    # ── PART 5 ─────────────────────────────────────────────
    ("PART 5\n배포 전략\n& 보안",
     "11강",
     "배포 플랫폼 10종 총정리",
     "- 10개 플랫폼 비교 실습 데모:\n"
     "  GitHub Pages / Streamlit / Vercel / Netlify / Firebase /\n"
     "  Railway / Render / Fly.io / HuggingFace Spaces / Replit\n"
     "- 무료·유료·제한사항·적합 프로젝트 유형별 정리\n"
     "- 실습: 기존 앱을 다른 플랫폼에 재배포",
     "30분"),
    (None,
     "12강",
     "환경 관리 & 보안 & CI/CD",
     "- .env 파일과 환경 변수 개념\n"
     "- .gitignore 설정 (절대 push하면 안 되는 파일)\n"
     "- GitHub Actions로 자동 배포 파이프라인 구축\n"
     "- 실습: .env + .gitignore + GitHub Actions 워크플로우 작성",
     "30분"),

    # ── PART 6 ─────────────────────────────────────────────
    ("PART 6\n텔레그램 봇\n+ 종합 실습",
     "13강",
     "텔레그램 봇 기초 + Claude API 연결",
     "- BotFather로 봇 토큰 발급\n"
     "- python-telegram-bot 라이브러리 구조\n"
     "- Claude API 연동 방법 (Polling vs Webhook)\n"
     "- 실습: 디스플레이 공정 상담 텔레그램 봇 제작 (bot.py)",
     "30분"),
    (None,
     "14강",
     "Gemini 봇 + 설비 알람 시스템",
     "- gemini_telebot 레포 분석 (waterfirst/gemini_telebot)\n"
     "- 알람 봇 설계 패턴 & Google Workspace 연동 개요\n"
     "- 실습: 수율 모니터링 알람 봇 구축\n"
     "  (임계값 초과 시 텔레그램 자동 알림)",
     "30분"),
    (None,
     "15강\n⏱️60분",
     "종합 통합 실습 (Streamlit + Gemini + 텔레그램)",
     "※ 이 강의만 60분\n"
     "- Claude에게 전체 아키텍처 설계 요청\n"
     "- Claude Code로 4개 파트 코드 생성:\n"
     "  data_simulator.py / app.py / alert_bot.py\n"
     "- 연결 테스트 & 디버깅 (강사 순회 지원)\n"
     "- GitHub push + Streamlit Cloud 배포\n"
     "※ 프로젝트 3 과제 안내 (수료 후 2주 이내 제출)",
     "60분"),
    (None,
     "16강",
     "최신 AI 트렌드 & 수료",
     "- 2025~2026 AI 개발 트렌드 (GitHub Top Trending 실시간 확인)\n"
     "- 디스플레이 × AI 실제 활용 사례 (삼성/LG 공개 사례)\n"
     "- 다음 학습 로드맵 추천\n"
     "- 프로젝트 3 결과 공유 및 수료\n"
     "※ 수료 조건: 프로젝트 1·2·3 모두 제출",
     "30분"),

    # ── PROJECT P3 ──────────────────────────────────────────
    ("PROJECT\n[P3]",
     "P3",
     "나만의 AI 디스플레이 도구 (자유 주제, 전체 스택)",
     "- 주제: 자유 주제 전체 스택 앱 제작\n"
     "- 하: 제공 풀스택 템플릿 커스터마이징\n"
     "- 중: Streamlit + Gemini/Claude API + 텔레그램 연결\n"
     "- 상: 주제 선정~아키텍처~구현~발표 전 과정 독립 수행\n"
     "- 발표: 3분 발표 + 배포 URL 제출\n"
     "※ 수료증 발급 기준 최종 과제",
     "2주일"),
]


# ══════════════════════════════════════════════════════════════
# ■ 데이터 행 기록
# ══════════════════════════════════════════════════════════════
START_ROW = 7
for i, (part, session, topic, detail, duration) in enumerate(rows):
    r = START_ROW + i
    ws.row_dimensions[r].height = 54.0

    is_proj = part is not None and part.startswith("PROJECT")
    row_bg = PROJ_BG if is_proj else PART_BG
    topic_bg = TOPIC_HL if is_proj else PART_BG

    # A열: 강의구분
    if part is not None:
        apply(ws.cell(r, 1), part,
              bold=True, bg=row_bg, halign="center")
    else:
        apply(ws.cell(r, 1), None, bg=PART_BG, halign="center")

    # B열: 차시
    apply(ws.cell(r, 2), session,
          bold=True, bg=row_bg, halign="center")

    # C열: 차시 주제
    apply(ws.cell(r, 3), topic,
          bold=True, bg=topic_bg, halign="left")

    # D열: 세부 내용
    apply(ws.cell(r, 4), detail,
          bold=False, bg=row_bg, halign="left")

    # E열: 예상시간
    apply(ws.cell(r, 5), duration,
          bold=False, bg=row_bg, halign="center")


# ══════════════════════════════════════════════════════════════
# ■ 합계 행 (양식 동일)
# ══════════════════════════════════════════════════════════════
total_row = START_ROW + len(rows)
ws.row_dimensions[total_row].height = 25.0
ws.merge_cells(f"A{total_row}:E{total_row}")
apply(ws.cell(total_row, 1),
      "총 강의 수: 16강  |  예상 총 학습시간: 약 8시간 30분 (강의) + 프로젝트 3종",
      bold=True, bg="FFBDD7EE", halign="center")


# ══════════════════════════════════════════════════════════════
# ■ 저장
# ══════════════════════════════════════════════════════════════
out = r"D:\python\2026_letuin\Letuin_AI_Lecture\docs\렛유인_KDC_강의계획서.xlsx"
wb.save(out)
print(f"저장 완료: {out}")
